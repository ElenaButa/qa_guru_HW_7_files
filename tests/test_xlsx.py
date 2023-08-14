from openpyxl import load_workbook
import os.path

# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_file():
    path_1 = (os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources/file_example_XLSX_50.xlsx'))
    workbook = load_workbook(path_1)
    sheet = workbook.active
    print(sheet.cell(row=2, column=3).value)
    assert (sheet.cell(row=3, column=2).value)=='Mara'
    assert (sheet.cell(row=6, column=2).value)=='Nereida'

